# -*- coding: utf-8 -*-
"""
文档解析服务：将各种格式的文档解析为文本块列表
支持：PDF（含书签）、Word(.docx)、Excel(.xlsx)、图片、Markdown/文本

优化：
- 表格行转自然语言（识别表头，行内容转"列名：值"句式）
- 多粒度分块（large 512字 + small 128字，两个粒度同时写入向量库）
- PDF 结构化解析（书签/TOC 作为章节标题）
- DOCX 标题样式识别（Heading 1/2/3 → 标题前缀）
"""
import logging
import os
from typing import Optional

logger = logging.getLogger(__name__)

# 分块参数
LARGE_CHUNK_SIZE = 512   # 大粒度：适合宏观问题
LARGE_CHUNK_OVERLAP = 80
SMALL_CHUNK_SIZE = 128   # 小粒度：适合精确匹配
SMALL_CHUNK_OVERLAP = 20

# 向下兼容
CHUNK_SIZE = LARGE_CHUNK_SIZE
CHUNK_OVERLAP = LARGE_CHUNK_OVERLAP


# ─────────────────────────────────────────────
# 通用文本分块
# ─────────────────────────────────────────────

def _split_text(text: str, chunk_size: int = LARGE_CHUNK_SIZE, overlap: int = LARGE_CHUNK_OVERLAP) -> list[str]:
    """将长文本按语义边界切分，带重叠"""
    text = text.strip()
    if not text:
        return []
    if len(text) <= chunk_size:
        return [text]
    chunks = []
    start = 0
    while start < len(text):
        end = start + chunk_size
        chunk = text[start:end]
        if end < len(text):
            for sep in ["\n\n", "。", "\n", ".", "；"]:
                idx = chunk.rfind(sep)
                if idx > chunk_size // 2:
                    chunk = chunk[:idx + len(sep)]
                    break
        chunks.append(chunk.strip())
        advance = max(1, len(chunk) - overlap)
        start += advance
    return [c for c in chunks if c]


def _split_multi_level(text: str, heading: str = "") -> list[dict]:
    """
    多粒度分块：同一文本生成 large（512字）和 small（128字）两个粒度。
    每个 chunk dict 包含 chunk_level 字段。
    heading 会作为前缀注入所有 chunk，保留上下文。
    """
    prefix = f"【{heading}】\n" if heading else ""
    results = []

    large_chunks = _split_text(text, LARGE_CHUNK_SIZE, LARGE_CHUNK_OVERLAP)
    for large_idx, large in enumerate(large_chunks):
        large_content = prefix + large
        results.append({
            "content": large_content,
            "chunk_level": "large",
            "parent_chunk_index": None,
        })
        small_chunks = _split_text(large, SMALL_CHUNK_SIZE, SMALL_CHUNK_OVERLAP)
        for small in small_chunks:
            results.append({
                "content": prefix + small,
                "chunk_level": "small",
                "parent_chunk_index": large_idx,
            })
    return results


# ─────────────────────────────────────────────
# 表格转自然语言
# ─────────────────────────────────────────────

def _table_to_nl(headers: list[str], rows: list[list[str]]) -> list[str]:
    """
    将表格转为自然语言句子列表。
    输入：
      headers = ["模块", "已落地基建内容", "状态"]
      rows    = [["Workmind测试平台", "需求任务管理...", "已上线"], ...]
    输出：
      ["模块：Workmind测试平台，已落地基建内容：需求任务管理...，状态：已上线。", ...]
    """
    sentences = []
    clean_headers = [h.strip() for h in headers if h.strip()]
    if not clean_headers:
        return sentences
    for row in rows:
        pairs = []
        for i, cell in enumerate(row):
            cell = cell.strip()
            if not cell:
                continue
            col = clean_headers[i] if i < len(clean_headers) else f"列{i+1}"
            pairs.append(f"{col}：{cell}")
        if pairs:
            sentences.append("，".join(pairs) + "。")
    return sentences


# ─────────────────────────────────────────────
# PDF 解析（含书签/TOC）
# ─────────────────────────────────────────────

def parse_pdf(file_path: str) -> list[dict]:
    """
    解析 PDF，使用 PyMuPDF。
    优先提取书签（TOC）作为章节标题，按章节分组分块。
    """
    try:
        import fitz
    except ImportError:
        logger.error("PyMuPDF 未安装，无法解析 PDF")
        return []

    chunks = []
    try:
        doc = fitz.open(file_path)
        toc = doc.get_toc()  # [[level, title, page], ...]

        # 建立 page → section_title 映射
        page_to_section: dict[int, str] = {}
        if toc:
            for i, (level, title, page) in enumerate(toc):
                next_page = toc[i + 1][2] if i + 1 < len(toc) else len(doc) + 1
                for p in range(page, next_page):
                    page_to_section[p] = title

        for page_idx in range(len(doc)):
            page = doc[page_idx]
            text = page.get_text("text").strip()
            has_image = len(page.get_images(full=False)) > 0
            page_num = page_idx + 1

            if not text:
                continue

            section = page_to_section.get(page_num, "")
            for item in _split_multi_level(text, heading=section):
                item.update({"page_num": page_num, "has_image": has_image, "image_path": None})
                chunks.append(item)

        doc.close()
    except Exception as e:
        logger.error(f"PDF 解析失败: {type(e).__name__}: {e!r}")
    return chunks


# ─────────────────────────────────────────────
# DOCX 解析（python-docx + 标题层级 + 表格转 NL）
# ─────────────────────────────────────────────

def parse_docx(file_path: str) -> list[dict]:
    """
    解析 Word 文档：
    - 用 python-docx 识别 Heading 1/2/3 样式作为章节标题
    - 表格识别表头行，数据行转自然语言
    - 多粒度分块
    """
    try:
        from docx import Document
        from docx.oxml.ns import qn
        from docx.text.paragraph import Paragraph
        from docx.table import Table
    except ImportError:
        logger.error("python-docx 未安装，无法解析 Word 文档")
        return []

    try:
        doc = Document(file_path)
    except Exception as e:
        logger.error(f"Word 文档打开失败: {type(e).__name__}: {e!r}")
        return []

    # 收集文档元素，保持原始顺序
    # (kind, data): kind ∈ {heading1, heading2, heading3, text, table_nl}
    elements: list[tuple[str, str]] = []

    body = doc.element.body
    for child in body.iterchildren():
        tag = child.tag

        if tag == qn("w:p"):
            para = Paragraph(child, doc)
            text = para.text.strip()
            if not text:
                continue
            style_name = para.style.name if para.style else ""
            if "Heading 1" in style_name or style_name in ("标题 1",):
                elements.append(("heading1", text))
            elif "Heading 2" in style_name or style_name in ("标题 2",):
                elements.append(("heading2", text))
            elif "Heading 3" in style_name or style_name in ("标题 3",):
                elements.append(("heading3", text))
            else:
                elements.append(("text", text))

        elif tag == qn("w:tbl"):
            table = Table(child, doc)
            rows_data: list[list[str]] = []
            for row in table.rows:
                cells = []
                seen = set()
                for cell in row.cells:
                    ct = cell.text.strip()
                    if ct and ct not in seen:
                        seen.add(ct)
                        cells.append(ct)
                if cells:
                    rows_data.append(cells)

            if not rows_data:
                continue

            # 第一行当表头，后续行转自然语言
            headers = rows_data[0]
            data_rows = rows_data[1:]
            if data_rows:
                nl_sentences = _table_to_nl(headers, data_rows)
                for s in nl_sentences:
                    elements.append(("table_nl", s))
            else:
                # 只有一行（纯表头或单行表格），直接当普通文本
                elements.append(("text", " | ".join(headers)))

    # 按标题分组，收集各章节的文本
    heading_stack: list[str] = []
    current_lines: list[str] = []
    sections: list[tuple[str, str]] = []  # [(heading_path, content_text)]

    def flush():
        if current_lines:
            path = " > ".join(heading_stack)
            sections.append((path, "\n".join(current_lines)))

    for kind, text in elements:
        if kind == "heading1":
            flush(); heading_stack = [text]; current_lines = []
        elif kind == "heading2":
            flush(); heading_stack = heading_stack[:1] + [text]; current_lines = []
        elif kind == "heading3":
            flush(); heading_stack = heading_stack[:2] + [text]; current_lines = []
        else:
            current_lines.append(text)
    flush()

    # 没有 Word 标题样式时，回退：把所有文字拼成一块再分
    if not sections:
        full_text = "\n".join(t for _, t in elements)
        sections = [("", full_text)] if full_text.strip() else []

    chunks = []
    for heading_path, content in sections:
        for item in _split_multi_level(content, heading=heading_path):
            item.update({"page_num": None, "has_image": False, "image_path": None})
            chunks.append(item)
    return chunks


# ─────────────────────────────────────────────
# XLSX 解析（表格转 NL）
# ─────────────────────────────────────────────

def parse_xlsx(file_path: str) -> list[dict]:
    """
    解析 Excel：第一行作为表头，数据行转自然语言句子。
    """
    try:
        import openpyxl
    except ImportError:
        logger.error("openpyxl 未安装，无法解析 Excel 文件")
        return []

    chunks = []
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows_data: list[list[str]] = []
            for row in ws.iter_rows(values_only=True):
                row_data = [str(cell).strip() if cell is not None else "" for cell in row]
                if any(c for c in row_data):
                    rows_data.append(row_data)

            if len(rows_data) < 2:
                continue

            headers = rows_data[0]
            data_rows = rows_data[1:]
            nl_sentences = _table_to_nl(headers, data_rows)

            heading = f"Sheet：{sheet_name}"
            content = "\n".join(nl_sentences)
            for item in _split_multi_level(content, heading=heading):
                item.update({"page_num": None, "has_image": False, "image_path": None})
                chunks.append(item)
        wb.close()
    except Exception as e:
        logger.error(f"Excel 解析失败: {type(e).__name__}: {e!r}")
    return chunks


# ─────────────────────────────────────────────
# 纯文本 / Markdown 解析
# ─────────────────────────────────────────────

def parse_text(file_path: str) -> list[dict]:
    """解析纯文本 / Markdown 文件，支持 # 标题识别"""
    chunks = []
    try:
        with open(file_path, "r", encoding="utf-8", errors="ignore") as f:
            content = f.read()

        # 按 Markdown 标题分段
        import re
        parts = re.split(r"(^#{1,3}\s+.+)$", content, flags=re.MULTILINE)
        heading = ""
        current_lines = []

        def flush_text():
            text = "\n".join(current_lines).strip()
            if text:
                for item in _split_multi_level(text, heading=heading):
                    item.update({"page_num": None, "has_image": False, "image_path": None})
                    chunks.append(item)

        for part in parts:
            part = part.strip()
            if not part:
                continue
            if re.match(r"^#{1,3}\s+", part):
                flush_text()
                heading = part.lstrip("#").strip()
                current_lines = []
            else:
                current_lines.append(part)
        flush_text()

        if not chunks:
            for item in _split_multi_level(content):
                item.update({"page_num": None, "has_image": False, "image_path": None})
                chunks.append(item)

    except Exception as e:
        logger.error(f"文本文件解析失败: {e}")
    return chunks


# ─────────────────────────────────────────────
# 图片解析（Vision API）
# ─────────────────────────────────────────────

def parse_image_with_vision(file_path: str, api_key: str, base_url: str, model: str) -> list[dict]:
    """使用 Vision API 解析图片内容"""
    import base64
    try:
        from openai import OpenAI
        client = OpenAI(api_key=api_key, base_url=base_url)
        ext = os.path.splitext(file_path)[1].lower()
        mime_map = {".jpg": "image/jpeg", ".jpeg": "image/jpeg", ".png": "image/png",
                    ".gif": "image/gif", ".webp": "image/webp", ".bmp": "image/bmp"}
        mime_type = mime_map.get(ext, "image/png")
        with open(file_path, "rb") as f:
            image_data = base64.b64encode(f.read()).decode("utf-8")
        response = client.chat.completions.create(
            model=model,
            messages=[{"role": "user", "content": [
                {"type": "image_url", "image_url": {"url": f"data:{mime_type};base64,{image_data}"}},
                {"type": "text", "text": "请详细描述这张图片的内容，如果包含文字、表格、流程图、架构图等，请完整提取并说明其含义。"},
            ]}],
        )
        content = response.choices[0].message.content or ""
        return [{"content": content, "chunk_level": "large", "parent_chunk_index": None,
                 "page_num": None, "has_image": True, "image_path": file_path}]
    except Exception as e:
        logger.error(f"Vision 图片解析失败: {e}")
        return []


# ─────────────────────────────────────────────
# 文件类型检测 & 统一入口
# ─────────────────────────────────────────────

def detect_file_type(file_name: str) -> str:
    ext = os.path.splitext(file_name)[1].lower()
    mapping = {
        ".pdf": "pdf", ".docx": "docx", ".doc": "docx",
        ".xlsx": "xlsx", ".xls": "xlsx",
        ".md": "md", ".txt": "txt",
        ".jpg": "image", ".jpeg": "image", ".png": "image",
        ".gif": "image", ".webp": "image", ".bmp": "image",
    }
    return mapping.get(ext, "other")


def parse_document(file_path: str, file_type: str, api_key: str = "", base_url: str = "", model: str = "") -> list[dict]:
    """
    统一入口：根据文件类型调用对应解析器
    返回: [{page_num, content, has_image, image_path, chunk_level, parent_chunk_index}]
    chunk_level: "large" | "small"
    """
    if file_type == "pdf":
        return parse_pdf(file_path)
    elif file_type == "docx":
        return parse_docx(file_path)
    elif file_type == "xlsx":
        return parse_xlsx(file_path)
    elif file_type in ("md", "txt"):
        return parse_text(file_path)
    elif file_type == "image":
        if api_key and base_url and model:
            return parse_image_with_vision(file_path, api_key, base_url, model)
        logger.warning("图片解析需要配置 AI Vision API Key")
        return []
    else:
        return parse_text(file_path)
